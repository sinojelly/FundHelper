# -*- coding: utf-8 -*-

import openpyxl
from FastFund import FastFund
from openpyxl.styles.numbers import FORMAT_NUMBER_00

from FundSheet import FUND_SHEET_NAME, UNIT_WORTH_COLUMN, CURRENT_PRICE_COLUMN, clear_sheet_columns


class FastFundSheet(object):
    def __init__(self, wb):
        self.work_book = wb
        self.sheet = wb[FUND_SHEET_NAME]
        # self.update_funds(None)

    def update_funds(self, invest_funds, progress_updater):
        row = 2
        for col in self.sheet.iter_cols(min_row=row, max_col=1):
            for cell in col:
                if cell.value is None:   # 遇到空行(id为空)，则后面不再更新
                    break

                fund_id = str(cell.value)
                fund = FastFund(fund_id)
                if not fund.initialize():
                    progress_updater()
                    row = row + 1
                    continue

                fixed_info_column_start = UNIT_WORTH_COLUMN  # left are fixed info
                clear_sheet_columns(self.sheet, row, fixed_info_column_start, 80)  # 把80列清空，目前表格模板够用且留有余量

                self.sheet['B' + str(row)].value = fund.fund_name
                self.sheet['B' + str(row)].hyperlink = "http://fund.eastmoney.com/{}.html".format(fund_id)

                self.sheet.cell(column=fixed_info_column_start, row=row).value = fund.unit_worth
                self.sheet.cell(column=fixed_info_column_start + 1, row=row).value = fund.unit_worth_change_ratio
                self.sheet.cell(column=fixed_info_column_start + 1, row=row).number_format = FORMAT_NUMBER_00
                self.sheet.cell(column=fixed_info_column_start + 5, row=row).value = fund.unit_worth_time # datetime.datetime.now()

                progress_updater()
                row = row + 1
            else:
                # Continue if the inner loop wasn't broken.
                continue
            # Inner loop was broken, break the outer.
            break

    def get_current_price(self, item_id):
        for col in self.sheet.iter_cols(min_row=2, max_col=1):
            for cell in col:
                if str(cell.value) == str(item_id):
                    return self.sheet.cell(row=cell.row, column=CURRENT_PRICE_COLUMN).value
        return None

    def get_sheet_name(self):
        return FUND_SHEET_NAME

    def get_invest_price(self, fund_id, time_str):
        return None, None

    def get_row_count(self):
        return self.sheet.max_row - 1  # 表头不用处理


if __name__ == '__main__':
    from TestTools import empty_func
    wb = openpyxl.load_workbook('test_model.xlsx')
    sheet = FastFundSheet(wb)
    sheet.update_funds([], empty_func)
    wb.save('test_model.xlsx')
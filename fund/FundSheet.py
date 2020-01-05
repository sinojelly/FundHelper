# -*- coding: utf-8 -*-

import openpyxl
from Fund import Fund, timestamp2time
from openpyxl.styles import Border, Side
from openpyxl.styles.colors import Color, BLUE, RED
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from XslxTools import get_cell_value, set_row_data, find_value_row_index, delete_rows, insert_row, \
    calc_change_ratio, update_focus_level, mark_name_delete

FUND_SHEET_NAME = "基金"

CURRENT_PRICE_COLUMN = 9
UNIT_WORTH_COLUMN = CURRENT_PRICE_COLUMN
HISTORY_WORTH_COLUMN_START = UNIT_WORTH_COLUMN + 12

CURRENT_PRICE_CHANGE_COLUMN = CURRENT_PRICE_COLUMN + 1

UNIT_WORTH_HISTORY_COLUMN = 4
FUND_FOCUS_LEVEL_COLUMN = 8

WEB_SHOW_COLUMNS = [1, 2, 4, 8, 9, 10, 11, 12, 14, 15, 17, 18, 20]


def clear_sheet_columns(work_sheet, row, column_start, column_num):
    # name 不清空，有些基金获取名称会失败，可能是手动写入的。对于自动获取的，后面会更新。
    # work_sheet.cell(column=2, row=row).value = None    # B列, name清空
    # work_sheet.cell(column=2, row=row).hyperlink = None

    # 清除6列自动填充数据
    for index in range(6):
        work_sheet.cell(column=column_start + index, row=row).value = None
        work_sheet.cell(column=column_start + index, row=row).border = None

    # 跳过6列手动维护的目标低值/高值
    column_start = HISTORY_WORTH_COLUMN_START
    for index in range(column_num):
        work_sheet.cell(column=column_start + index, row=row).value = None
        work_sheet.cell(column=column_start + index, row=row).border = None


def calc_ac_worth_continuous_change(fund):
    days = fund.continuous_days
    if days >= len(fund.ac_worth_trend):
        print("calc_ac_worth_continuous_change days =", days, "ac_worth_trend.length =", fund.ac_worth_trend.length)
        return 0
    old_value = fund.ac_worth_trend[-days-1][1]
    ratio = calc_change_ratio(old_value, fund.ac_worth)
    # print("calc_ac_worth_continuous_change old_value", old_value)
    # print("ac worth trend:", fund.ac_worth_trend[-days-2:])
    # print("ratio:", ratio, "days:", days)
    return ratio


def get_fund_hyperlink(fund_id):
    return "http://fund.eastmoney.com/{}.html".format(fund_id)


class FundSheet(object):
    def __init__(self, wb):
        self.work_book = wb
        self.sheet = wb[FUND_SHEET_NAME]
        self.invested_funds_map = {}
        # self.update_funds()

    def update_funds(self, invested_funds, progress_updater):
        row = 2
        for col in self.sheet.iter_cols(min_row=row, max_col=1):
            for cell in col:
                if cell.value is None:   # 遇到空行(id为空)，则后面不再更新
                    break

                fixed_info_column_start = UNIT_WORTH_COLUMN  # left are fixed info
                clear_sheet_columns(self.sheet, row, fixed_info_column_start, 80)  # 把80列清空，目前表格模板够用且留有余量

                fund_id = str(cell.value)
                fund = Fund(fund_id)
                if fund.initialize():
                    current_fund_buy = False
                    if fund_id in invested_funds:
                        self.invested_funds_map[fund_id] = fund
                        current_fund_buy = True
                        # print("store invested fund", fund_id)

                    self.sheet['B' + str(row)].value = fund.fund_name
                    self.sheet['B' + str(row)].hyperlink = get_fund_hyperlink(fund_id)

                    if fund.unit_worth_history is not None:
                        unit_worth_history_str = str(fund.unit_worth_history)
                        unit_worth_history_str = unit_worth_history_str[1:-1]   # 去掉中括号
                        self.sheet.cell(column=UNIT_WORTH_HISTORY_COLUMN, row=row).value = unit_worth_history_str
                    update_focus_level(self.sheet, FUND_FOCUS_LEVEL_COLUMN, row, current_fund_buy)

                    self.sheet.cell(column=fixed_info_column_start, row=row).value = fund.unit_worth
                    self.sheet.cell(column=fixed_info_column_start + 1, row=row).value = fund.unit_worth_change_ratio
                    self.sheet.cell(column=fixed_info_column_start + 1, row=row).number_format = FORMAT_NUMBER_00
                    self.sheet.cell(column=fixed_info_column_start + 2, row=row).value = fund.continuous_days
                    self.sheet.cell(column=fixed_info_column_start + 3, row=row).value = calc_ac_worth_continuous_change(fund)
                    self.sheet.cell(column=fixed_info_column_start + 4, row=row).value = fund.ac_worth
                    self.sheet.cell(column=fixed_info_column_start + 5, row=row).value = fund.unit_worth_time # datetime.datetime.now()

                    self.update_target_values(fund, row, fixed_info_column_start + 6)
                    # self.update_history_worth(fund, row)

                progress_updater()   # 每次row增加前+1
                row = row + 1
            else:
                # Continue if the inner loop wasn't broken.
                continue
            # Inner loop was broken, break the outer.
            break

    # 涉及目标低值/高值，从start_column开始的6列
    def update_target_values(self, fund, row, start_column):
        self.set_target_price(fund, row, start_column + 2)
        self.set_target_price(fund, row, start_column + 5)

    def set_target_price(self, fund, row, time_column):
        target_price_cell = self.sheet.cell(column=time_column - 1, row=row)
        str_target_price = get_cell_value(self.sheet, target_price_cell)
        if str_target_price != '':   # 已经有值
            return
        str_date = get_cell_value(self.sheet, self.sheet.cell(column=time_column, row=row))
        if str_date != '':
            target_price = fund.get_price(str_date)
            if target_price is not None:
                target_price_cell.value = target_price

    def update_history_worth(self, fund, row):
        column_start = HISTORY_WORTH_COLUMN_START
        # print("fund.recent_ac_worth_max")
        # print(fund.recent_ac_worth_max)
        min_min_value = 99999  # 极小值中最小的一个
        min_min_column = 0  # 极小值中最小的一个对应的列
        max_max_value = 0  # 极大值中最大的一个
        max_max_column = 0  # 极大值中最小的一个对应的列
        ac_worth_max = list(fund.recent_ac_worth_max[::-1])
        for min_item in fund.recent_ac_worth_min[::-1]:
            max_item = None
            if len(ac_worth_max) > 0:
                max_item = ac_worth_max.pop(0)
            count, min_min_value, min_min_column, max_max_value, max_max_column = self.write_extrema_worth(column_start,
                                                                                                           row,
                                                                                                           min_item,
                                                                                                           max_item,
                                                                                                           fund.ac_worth,
                                                                                                           min_min_value,
                                                                                                           min_min_column,
                                                                                                           max_max_value,
                                                                                                           max_max_column)
            column_start = column_start + count
        side = Side(border_style='medium', color=Color(rgb=BLUE))
        border = Border(left=side, right=side, top=side, bottom=side)
        self.sheet.cell(column=min_min_column, row=row).border = border
        side = Side(border_style='medium', color=Color(rgb=RED))
        border = Border(left=side, right=side, top=side, bottom=side)
        self.sheet.cell(column=max_max_column, row=row).border = border

    def write_extrema_worth(self, column_start, row, min_data, max_data, current, min_value, min_column, max_value, max_column):
        ratio = (min_data[1] - current) / current
        self.sheet.cell(column=column_start, row=row).border = None  # 清除原来的框
        self.sheet.cell(column=column_start, row=row).value = ratio * 100   # 换算成百分数
        self.sheet.cell(column=column_start, row=row).number_format = FORMAT_NUMBER_00
        self.sheet.cell(column=column_start + 1, row=row).value = min_data[1]
        self.sheet.cell(column=column_start + 2, row=row).value = timestamp2time(min_data[0])
        if min_data[1] < min_value:
            min_value = min_data[1]
            min_column = column_start
        if max_data is not None:
            ratio = (max_data[1] - current) / current
            self.sheet.cell(column=column_start + 3, row=row).border = None
            self.sheet.cell(column=column_start + 3, row=row).value = ratio * 100   # 换算成百分数
            self.sheet.cell(column=column_start + 3, row=row).number_format = FORMAT_NUMBER_00
            self.sheet.cell(column=column_start + 4, row=row).value = max_data[1]
            self.sheet.cell(column=column_start + 5, row=row).value = timestamp2time(max_data[0])
            if max_data[1] > max_value:
                max_value = max_data[1]
                max_column = column_start + 3
            return 6,min_value,min_column,max_value,max_column
        return 6,min_value,min_column,max_value,max_column  # 把3改成6，无论是否有H的值，都把它的列空出来，避免H的列写L的值

    def get_current_price(self, item_id):
        for col in self.sheet.iter_cols(min_row=2, max_col=1):
            for cell in col:
                if str(cell.value) == str(item_id):
                    return self.sheet.cell(row=cell.row, column=CURRENT_PRICE_COLUMN).value, self.sheet.cell(row=cell.row, column=CURRENT_PRICE_CHANGE_COLUMN).value
        return None

    def get_sheet_name(self):
        return FUND_SHEET_NAME

    def get_invest_price(self, fund_id, time_str):
        fund = self.invested_funds_map[str(fund_id)]
        if fund is None:
            return None
        return fund.get_price(time_str), fund.fund_name

    def get_row_count(self):
        return self.sheet.max_row - 1   # 表头去掉

    def get_table(self):
        result = []
        row_index = 2
        for row in self.sheet.iter_rows(min_row=row_index, max_col=1):
            for cell in row:
                if cell.value is None:  # 遇到空行(id为空)，则后面不再更新
                    break
                row_result = []
                for column in WEB_SHOW_COLUMNS:
                    cell = self.sheet.cell(row=row_index, column=column)
                    # value = reader.get_cell_value(cell.coordinate, FUND_SHEET_NAME)
                    value = get_cell_value(self.sheet, cell)
                    row_result.append(value)
                    if column == 2:  # name
                        fund_id = self.sheet.cell(row=row_index, column=1).value
                        link = get_fund_hyperlink(fund_id)
                        row_result.append(link)
                result.append(row_result)
                row_index = row_index + 1
            else:
                # Continue if the inner loop wasn't broken.
                continue
            # Inner loop was broken, break the outer.
            break
        return result

    def save_table(self, data):
        # 先把所有行第二列（名称）改为delete，然后根据web传来的数据覆盖excel内容，最后把名称为delete的删掉
        mark_name_delete(self.sheet, 2, 2)
        for row_data in data:   # for each row data
            row_index = find_value_row_index(self.sheet, 2, 1, row_data[0])
            # print('row data:', row_data[0], row_data[1], ", row_index=", row_index)
            if row_index is None:   # 原来不存在添加行
                del row_data[2]    # 删除 hyperlinnk
                insert_row(self.sheet, 2, 1, row_data, WEB_SHOW_COLUMNS)
            else:
                del row_data[2]    # 删除 hyperlinnk
                set_row_data(self.sheet, row_index, row_data, 2, WEB_SHOW_COLUMNS)
        delete_rows(self.sheet, 2, 2)


if __name__ == '__main__':
    from TestTools import empty_func
    wb = openpyxl.load_workbook('test_model - 副本.xlsx')
    sheet = FundSheet(wb)
    # sheet.update_funds([], empty_func)
    tabledata = [
        ["005911", "广发双擎升级混合AAA", "6", "12.0368", "0.17", "3", "20191210", "-15.21", "20191201", "-1.02", "20191202"],
        ["519674", "诺安成长混合", "4", "1.195", "0.5", "4", "20191210", "-16.15", "20191021", "-3.77", "20191119"],
        ["005999", "广发双擎升级混合CCC", "6", "12.0368", "0.17", "5", "20191210", "-15.21", "20191201", "-1.02", "20191202"],
        ["320088", "诺安成长混合BB", "4", "1.195", "0.5", "4", "20191210", "-16.15", "20191021", "-3.77", "20191119"]]
    sheet.save_table(tabledata)
    wb.save('test_model - 副本.xlsx')
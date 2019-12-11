# -*- coding: utf-8 -*-
import datetime

from openpyxl import load_workbook
from io import BytesIO

from FastFundSheet import FastFundSheet
from FundSheet import FUND_SHEET_NAME, FundSheet
from InvestSheet import InvestSheet
from StockIndexSheet import StockIndexSheet


class FundWorkbook(object):
    def __init__(self, file):
        self.file = file
        print("load workbook: ", file)
        self.wb = load_workbook(file)   # self.reader.get_book()
        self.sheet = {FUND_SHEET_NAME: FundSheet(self.wb)}  # 注意不是原始wb 中的Sheet，而是这里的FundSheet包装类对象
        self.sheet_alias = {FUND_SHEET_NAME: 'fund'}  # 使用英文别名，便于在javascript中直接访问
        self.alias_to_sheet = {}
        for key,value in self.sheet_alias.items():   # 根据前面配置自动配置sheet_alias 到 sheet对象映射表，便于数据更新
            self.alias_to_sheet[value] = self.sheet[key]

    def get_table(self, sheet_name=None):
        return {self.sheet_alias[FUND_SHEET_NAME]: self.sheet[FUND_SHEET_NAME].get_table()}

    def save_table(self, data):
        for key,value in data.items():
            # 每个key对应一个sheet
            # print("save_table key is ", key)
            self.alias_to_sheet[key].save_table(value)
        self.wb.save(self.file)

    def update_funds(self, fast_run, progress_updater):
        print("Start ...")
        starttime = datetime.datetime.now()

        invest_sheet = InvestSheet(self.wb)
        invest_funds = invest_sheet.get_all_funds()

        if fast_run:
            fund_sheet = FastFundSheet(self.wb)
        else:
            fund_sheet = FundSheet(self.wb)
        stock_index_sheet = StockIndexSheet(self.wb)

        fund_sheet_row_count = fund_sheet.get_row_count()
        stock_index_sheet_row_count = stock_index_sheet.get_row_count()
        step_count = fund_sheet_row_count + stock_index_sheet_row_count + 1  # invest sheet更新作为1步

        progress_updater(total=step_count)  # 更新步骤总数

        fund_sheet.update_funds(invest_funds, progress_updater)

        end_funds_time = datetime.datetime.now()
        print("Finished update funds! It takes", (end_funds_time - starttime).seconds, "seconds.")

        stock_index_sheet.update_stock_index(progress_updater)

        end_stock_index_time = datetime.datetime.now()

        print("Finished update stock index! It takes", (end_stock_index_time - end_funds_time).seconds, "seconds.")

        invest_sheet.update_all_invests(fund_sheet, stock_index_sheet)
        # wb.save(file)  # not save to the origin file

        progress_updater(finished=True)  # 更新当前步骤，并且表明已经结束

        endtime = datetime.datetime.now()
        print("Finished all! It takes", (endtime - starttime).seconds, "seconds.")

    def download_excel(self):
        virtual_workbook = BytesIO()
        self.wb.save(virtual_workbook)

        return virtual_workbook.getvalue()  # 返回临时生成的文件内容，而不是修改文件。避免多应用访问把文件改乱



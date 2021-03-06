# -*- coding: utf-8 -*-

import openpyxl

from openpyxl.styles.numbers import FORMAT_NUMBER_00

from XslxTools import set_p_n_condition, str_to_float, is_value_empty, set_double_p_n_condition
from XslxTools import get_cell_value, set_row_data, find_value_row_index, get_index_range, insert_row

import logging
_logger = logging.getLogger('werkzeug')

INVEST_SHEET_NAME = "投资"

INVEST_COLUMN_START = 4
INVEST_ROW_START = 6


class InvestSheet(object):
    def __init__(self, wb):
        self.work_book = wb
        self.sheet = wb[INVEST_SHEET_NAME]
        self.stock_sheets = {}

    def get_price_sheet(self, col):
        price_sheet_name = self.sheet.cell(column=col, row=1).value
        # print("col", col)
        # print("price_sheet_name", price_sheet_name)
        price_sheet = self.stock_sheets[price_sheet_name]
        return price_sheet

    def get_fund_or_stockindex_id(self, col):
        fund_id = self.sheet.cell(column=col, row=2).value
        # print("get_fund_or_stockindex_id", fund_id)
        return fund_id

    def get_invest_price(self, row, column):
        price_sheet = self.get_price_sheet(column)
        invest_time = str(self.sheet.cell(row=row, column=2).value)
        _logger.info("invest time: " + invest_time + "row: " + str(row) + "col: " + str(column))
        invest_price, fund_name = price_sheet.get_invest_price(self.get_fund_id(column), invest_time)
        return invest_price, fund_name

    def get_current_price(self, col):
        price_sheet = self.get_price_sheet(col)

        item_id = self.get_fund_or_stockindex_id(col)
        price = None
        price_change = None
        try:
            price, price_change = price_sheet.get_current_price(item_id)
        except TypeError as err:
            _logger.error("price_sheet.get_current_price TypeError. fund/stock_id = "
                          + str(item_id) + ", col = " + str(col) + ", exception: " + str(err))

        if price is None:
            # print("get_current_price fail, col =", col, ", id =", item_id)
            _logger.error("get_current_price fail, col = " + str(col) + ", id = " + str(item_id))

        return price, price_change

    def get_invest_status(self, row):
        return self.sheet.cell(row=row, column=3).value

    # 处理每一个基金
    def process_one_fund(self, col):
        price, price_change = self.get_current_price(col)
        if price is None:
            return 0, 0, 0
        total_invest = 0
        total_income = 0
        need_process = True
        for col_content in self.sheet.iter_cols(min_col=col, max_col=col, min_row=INVEST_ROW_START):  # 只遍历基金ID那一列
            for cell in col_content:
                if not need_process:   # 跳过一行，再处理。因为每一组投资都是两行
                    need_process = True
                    continue
                need_process = False
                if is_value_empty(cell.value):  # 跳过未填写的
                    continue
                status = self.get_invest_status(cell.row)
                # print("row =", cell.row, ", status =", status)
                if status != '投资':
                    continue
                float_value = str_to_float(cell.value)
                if float_value is None:
                    continue
                invest_amount = float(float_value)

                invest_price = self.sheet.cell(column=int(cell.col_idx)+1, row=cell.row).value
                invest_price = str_to_float(invest_price)   # None or '' both need get_invest_price
                if invest_price is None:
                    invest_price, fund_name = self.get_invest_price(cell.row, cell.col_idx)
                    if is_value_empty(invest_price) is not True:
                        self.sheet.cell(column=int(cell.col_idx) + 1, row=cell.row).value = float(invest_price)  # 写入投资价格
                        self.sheet.cell(column=int(cell.col_idx) + 1, row=2).value = fund_name    # 写入基金名称(指数对应的组合名称不能自动写入)
                if is_value_empty(invest_price):  # 跳过查询不到投资价格的
                    _logger.info("error: invest_price none for fund_name = " + fund_name)
                    #print("error: invest_price none for fund_name = " + fund_name)
                    continue
                invest_price = float(invest_price)

                if invest_amount > 0:
                    total_invest = total_invest + invest_amount
                    ratio = (price - invest_price)/invest_price
                    income = invest_amount * ratio
                    total_income = total_income + income
                    income_cell = self.sheet.cell(column=int(cell.col_idx), row=cell.row + 1)
                    income_cell.value = income
                    income_cell.number_format = FORMAT_NUMBER_00
                    set_p_n_condition(self.sheet, income_cell)
                    ratio_cell = self.sheet.cell(column=int(cell.col_idx) + 1, row=cell.row + 1)
                    ratio_cell.value = ratio * 100  # ratio
                    ratio_cell.number_format = FORMAT_NUMBER_00
                    set_p_n_condition(self.sheet, ratio_cell)
                elif invest_amount < 0:    # 表示后面的投资都已经卖掉了
                    break
            else:
                # Continue if the inner loop wasn't broken.
                continue
            # Inner loop was broken, break the outer.
            break
            # https://stackoverflow.com/questions/189645/how-to-break-out-of-multiple-loops

        recent_day_change = 0
        if total_invest != 0:
            self.sheet.cell(column=col, row=3).value = total_invest    # 总投资额
            self.sheet.cell(column=col+1, row=3).value = price           # 当前价格/点数
            self.sheet.cell(column=col, row=4).value = total_income   # 总收益
            self.sheet.cell(column=col, row=4).number_format = FORMAT_NUMBER_00
            set_p_n_condition(self.sheet, self.sheet.cell(column=col, row=4))
            self.sheet.cell(column=col+1, row=4).value = total_income / total_invest * 100   # 总收益率
            self.sheet.cell(column=col+1, row=4).number_format = FORMAT_NUMBER_00
            set_p_n_condition(self.sheet, self.sheet.cell(column=col+1, row=4))

            # 最近一天变化情况
            if price_change is not None:
                recent_day_change = total_invest * price_change / 100
                recent_day_change_cell = self.sheet.cell(column=col, row=5)
                recent_day_change_cell.value = recent_day_change  # 最近一天收益
                set_double_p_n_condition(self.sheet, recent_day_change_cell)
                recent_day_change_ratio_cell = self.sheet.cell(column=col + 1, row=5)
                recent_day_change_ratio_cell.value = price_change  # 最近一天收益率
                set_double_p_n_condition(self.sheet, recent_day_change_ratio_cell)
        return total_invest, total_income, recent_day_change

    def update_all_invests(self, *stock_sheets):
        for stock_sheet in stock_sheets:
            self.stock_sheets[stock_sheet.get_sheet_name()] = stock_sheet

        total_invest = 0
        total_income = 0
        total_recent_change = 0
        need_process = True
        for col in self.sheet.iter_cols(min_col=INVEST_COLUMN_START, min_row=2, max_row=2):   # 只读基金/指数ID
            for cell in col:
                if not need_process:  # 跳过一列，再处理。因为每一个基金都是两列
                    need_process = True
                    continue
                need_process = False
                if cell.value is None:  # 跳过未填写的
                    break
                invest, income, recent_change = self.process_one_fund(int(cell.col_idx))
                total_invest += invest
                total_income += income
                total_recent_change += recent_change
        self.sheet.cell(row=3, column=2).value = total_invest
        self.sheet.cell(row=4, column=2).value = total_income
        if total_invest != 0:
            self.sheet.cell(row=4, column=3).value = total_income/total_invest*100
        self.sheet.cell(row=4, column=2).number_format = FORMAT_NUMBER_00
        set_p_n_condition(self.sheet, self.sheet.cell(row=4, column=2))
        self.sheet.cell(row=4, column=3).number_format = FORMAT_NUMBER_00
        set_p_n_condition(self.sheet, self.sheet.cell(row=4, column=3))

        # 最近一天变化
        recent_change_cell = self.sheet.cell(row=5, column=2)
        recent_change_cell.value = total_recent_change
        set_double_p_n_condition(self.sheet, recent_change_cell)
        if total_invest != 0:
            recent_change_ratio_cell = self.sheet.cell(row=5, column=3)
            recent_change_ratio_cell.value = total_recent_change/total_invest*100
            set_double_p_n_condition(self.sheet, recent_change_ratio_cell)

    def get_fund_id(self, col):
        category = self.sheet.cell(row=1, column=col).value
        if category is not None and category == "基金":
            return self.sheet.cell(row=2, column=col).value
        return None

    # 获取投资过的基金列表（便于查询历史价格），以及仍然在投资中的基金（便于标注focus level）
    # 暂时只使用了 still_invest_funds, 因为它可以支持上述两种场景，更省内存
    def get_all_funds(self):
        invested_funds = []
        still_invest_funds = []
        need_process = True
        for col in self.sheet.iter_cols(min_col=INVEST_COLUMN_START, min_row=1, max_row=1):
            for cell in col:
                if not need_process:  # 跳过一行，再处理。因为每一组投资都是两行
                    need_process = True
                    continue
                need_process = False
                if cell.value is None:  # 跳过未填写的
                    break
                fund_id = self.get_fund_id(int(cell.col_idx))
                if fund_id is not None:
                    invested_funds.append(str(fund_id))
                    if self.is_still_invest(int(cell.col_idx)):
                        still_invest_funds.append(str(fund_id))
        return invested_funds, still_invest_funds

    # 仍然未完全赎回
    def is_still_invest(self, col_id):
        need_process = True
        for col_content in self.sheet.iter_cols(min_col=col_id, max_col=col_id, min_row=INVEST_ROW_START):  # 只遍历基金ID那一列
            for cell in col_content:
                if not need_process:  # 跳过一行，再处理。因为每一组投资都是两行
                    need_process = True
                    continue
                need_process = False
                if is_value_empty(cell.value):  # 跳过未填写的
                    continue
                status = self.get_invest_status(cell.row)
                # print("row =", cell.row, ", status =", status)
                if status == '投资':
                    return True
        return False

    def get_table(self):
        result = []
        row_index = 1
        for row in self.sheet.iter_rows(min_row=row_index, max_col=1):
            for cell in row:
                row_result = []
                for column in get_index_range(self.sheet.max_column):
                    cell = self.sheet.cell(row=row_index, column=column)
                    value = get_cell_value(self.sheet, cell)
                    row_result.append(value)
                result.append(row_result)
                row_index = row_index + 1
        # print("InvestSheet get table : length = ", len(result[0]))
        return result

    def save_table(self, data):
        for row_data in data:   # for each row data
            # print("InvestSheet save table : length = ", len(row_data))
            row_index = find_value_row_index(self.sheet, 1, 1, row_data[0])
            if row_index is None:   # 原来不存在添加行
                insert_row(self.sheet, 1, 1, row_data)
            else:
                set_row_data(self.sheet, row_index, row_data, 1)


if __name__ == '__main__':
    wb = openpyxl.load_workbook('Jelly_model.xlsx')
    import StockIndexSheet
    stock_index_sheet = StockIndexSheet.StockIndexSheet(wb)
    import FundSheet
    fund_sheet = FundSheet.FundSheet(wb)
    invest_sheet = InvestSheet(wb)
    invest_funds, still_invested_funds = invest_sheet.get_all_funds()
    from TestTools import empty_func
    fund_sheet.update_funds(still_invested_funds, empty_func)
    invest_sheet.update_all_invests(fund_sheet, stock_index_sheet)
    wb.save('Jelly_model_new.xlsx')

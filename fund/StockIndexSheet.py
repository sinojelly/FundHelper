# -*- coding: utf-8 -*-

import openpyxl
from StockIndex import StockIndex
from EastMoneyPushStockIndex import EastMoneyPushStockIndex
from JuheStockIndex import JuheStockIndex

from openpyxl.styles import Border, Side
from openpyxl.styles.colors import Color, BLUE, RED
from openpyxl.styles.numbers import FORMAT_NUMBER_00

from XslxTools import get_cell_value, find_value_row_index, insert_row, \
    set_row_data, delete_rows, update_focus_level, mark_name_delete


STOCK_SHEET_NAME = "指数"

CURRENT_INDEX_COLUMN = 6

CURRENT_INDEX_CHANGE_COLUMN = CURRENT_INDEX_COLUMN + 1

RELATED_FUND_IDS_COLUMN = 4
FOCUS_LEVEL_COLUMN = 5

# excel: start from 1,  javascript: start from 0
WEB_SHOW_COLUMNS = [1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]

STOCK_LINK_COLUMN = 3

DEFAULT_STOCK_LINK = "https://finance.sina.com.cn"


def clear_sheet_columns(work_sheet, row, column_start, column_num):
    # 当前值相关信息清空
    work_sheet.cell(column=6, row=row).value = None    # 点数清空
    work_sheet.cell(column=7, row=row).hyperlink = None
    work_sheet.cell(column=8, row=row).hyperlink = None
    work_sheet.cell(column=9, row=row).hyperlink = None
    work_sheet.cell(column=10, row=row).hyperlink = None

    # 目标比例清空
    work_sheet.cell(column=11, row=row).hyperlink = None
    work_sheet.cell(column=14, row=row).hyperlink = None

    # 自动生成极大极小值相关信息清空
    for index in range(column_num):
        work_sheet.cell(column=column_start + index, row=row).value = None
        work_sheet.cell(column=column_start + index, row=row).border = None


class StockIndexSheet(object):
    def __init__(self, wb):
        self.work_book = wb
        self.sheet = wb[STOCK_SHEET_NAME]
        # self.update_stock_index()

    def update_focus_level(self, row, stock_id, invested_funds):
        current_fund_buy = False
        if stock_id in invested_funds:  # 如果购买的是指数，或者股票，直接判断
            current_fund_buy = True
        else:  # 如果购买的是指数关联的基金，则看关联基金是否购买了
            fund_id_str = get_cell_value(self.sheet, self.sheet.cell(column=RELATED_FUND_IDS_COLUMN, row=row))
            fund_ids = str(fund_id_str).split(',')  # 逗号分隔的多个FundID
            for fund_id in fund_ids:  # 任何一个Fund 购买了，都认为该指数购买了
                if fund_id in invested_funds:
                    current_fund_buy = True
                    break
        update_focus_level(self.sheet, FOCUS_LEVEL_COLUMN, row, current_fund_buy)

    def update_stock_index(self, invested_funds, progress_updater):
        row = 2
        for col in self.sheet.iter_cols(min_row=row, max_col=1):
            for cell in col:
                fixed_info_column_start = 6
                auto_extrema_column_start = fixed_info_column_start+12
                clear_sheet_columns(self.sheet, row, auto_extrema_column_start, 80)  # 把80列清空，目前表格模板够用且留有余量

                stock_index_id = str(cell.value)
                self.update_focus_level(cell.row, stock_index_id, invested_funds)
                stock_index = StockIndex(stock_index_id)
                if not stock_index.initialize():
                    eastmoney_stock_index = EastMoneyPushStockIndex(stock_index_id)
                    if eastmoney_stock_index.initialize():
                        self.sheet.cell(column=fixed_info_column_start, row=row).value = eastmoney_stock_index.current_index
                        self.sheet.cell(column=fixed_info_column_start + 1,
                                        row=row).value = eastmoney_stock_index.current_index_change_ratio
                        self.sheet.cell(column=fixed_info_column_start + 5,
                                        row=row).value = eastmoney_stock_index.current_index_time
                    else:
                        juhe_stock_index = JuheStockIndex(stock_index_id)
                        if juhe_stock_index.initialize():
                            self.sheet.cell(column=fixed_info_column_start,
                                            row=row).value = juhe_stock_index.current_index
                            self.sheet.cell(column=fixed_info_column_start + 1,
                                            row=row).value = juhe_stock_index.current_index_change_ratio
                            self.sheet.cell(column=fixed_info_column_start + 5,
                                            row=row).value = juhe_stock_index.current_index_time
                    progress_updater()
                    row = row + 1
                    continue
                # self.sheet['B' + str(row)].value = stock_index.fund_name
                # self.sheet['B' + str(row)].hyperlink = "http://fund.eastmoney.com/{}.html".format(stock_index_id)

                self.sheet.cell(column=fixed_info_column_start, row=row).value = stock_index.current_index
                self.sheet.cell(column=fixed_info_column_start + 1, row=row).value = stock_index.current_index_change_ratio
                self.sheet.cell(column=fixed_info_column_start + 1, row=row).number_format = FORMAT_NUMBER_00
                self.sheet.cell(column=fixed_info_column_start + 2, row=row).value = stock_index.continuous_days
                self.sheet.cell(column=fixed_info_column_start + 3, row=row).value = stock_index.continuous_ratio
                self.sheet.cell(column=fixed_info_column_start + 3, row=row).number_format = FORMAT_NUMBER_00
                self.sheet.cell(column=fixed_info_column_start + 4, row=row).value = stock_index.current_amount
                self.sheet.cell(column=fixed_info_column_start + 4, row=row).number_format = FORMAT_NUMBER_00
                self.sheet.cell(column=fixed_info_column_start + 5, row=row).value = stock_index.current_index_time

                if stock_index.recent_index_max is None:
                    progress_updater()
                    row = row + 1
                    continue

                # self.update_history_worth(stock_index, row, auto_extrema_column_start)

                progress_updater()
                row = row + 1

    def update_history_worth(self, stock_index, row, auto_extrema_column_start):
        # 后面是自动生成的极大极小值比例信息
        column_start = auto_extrema_column_start
        # print("fund.recent_ac_worth_max")
        # print(fund.recent_ac_worth_max)
        min_min_value = 99999  # 极小值中最小的一个
        min_min_column = 0  # 极小值中最小的一个对应的列
        max_max_value = 0  # 极大值中最大的一个
        max_max_column = 0  # 极大值中最小的一个对应的列
        index_max = list(stock_index.recent_index_max)
        for min_item in stock_index.recent_index_min:
            max_item = None
            if len(index_max) > 0:
                max_item = index_max.pop(0)
            count, min_min_value, min_min_column, max_max_value, max_max_column = self.write_extrema_index(column_start,
                                                                                                           row,
                                                                                                           min_item,
                                                                                                           max_item,
                                                                                                           stock_index.current_index,
                                                                                                           min_min_value,
                                                                                                           min_min_column,
                                                                                                           max_max_value,
                                                                                                           max_max_column)
            column_start = column_start + count
        side = Side(border_style='medium', color=Color(rgb=BLUE))
        border = Border(left=side, right=side, top=side, bottom=side)
        self.sheet.cell(column=min_min_column, row=row).border = border
        side = Side(border_style='medium', color=Color(rgb=RED))
        border = Border(left=side, right=side, top=side, bottom=side)
        self.sheet.cell(column=max_max_column, row=row).border = border

    def write_extrema_index(self, column_start, row, min_data, max_data, current, min_value, min_column, max_value, max_column):
        ratio = (min_data[1] - current) / current
        self.sheet.cell(column=column_start, row=row).border = None  # 清除原来的框
        self.sheet.cell(column=column_start, row=row).value = ratio * 100   # 换算成百分数
        self.sheet.cell(column=column_start + 1, row=row).value = min_data[1]
        self.sheet.cell(column=column_start + 2, row=row).value = min_data[0]
        if min_data[1] < min_value:
            min_value = min_data[1]
            min_column = column_start
        if max_data is not None:
            ratio = (max_data[1] - current) / current
            self.sheet.cell(column=column_start + 3, row=row).border = None
            self.sheet.cell(column=column_start + 3, row=row).value = ratio * 100   # 换算成百分数
            self.sheet.cell(column=column_start + 4, row=row).value = max_data[1]
            self.sheet.cell(column=column_start + 5, row=row).value = max_data[0]
            if max_data[1] > max_value:
                max_value = max_data[1]
                max_column = column_start + 3
            return 6,min_value,min_column,max_value,max_column
        return 6,min_value,min_column,max_value,max_column  # 把3改成6，无论是否有H的值，都把它的列空出来，避免H的列写L的值

    def get_current_price(self, item_id):
        for col in self.sheet.iter_cols(min_row=2, max_col=1):
            for cell in col:
                if str(cell.value) == str(item_id):
                    return self.sheet.cell(row=int(cell.row), column=CURRENT_INDEX_COLUMN).value, self.sheet.cell(row=int(cell.row), column=CURRENT_INDEX_CHANGE_COLUMN).value
        return None

    def get_sheet_name(self):
        return STOCK_SHEET_NAME

    def get_row_count(self):
        return self.sheet.max_row - 1  # 去掉表头

    # 目前没有历史信息，不能查到指定某天的价格，所以直接返回空。Invest表需要手动填写名称和投资时价格点数
    def get_invest_price(self, fund_id, time_str):
        return None

    def get_table(self):
        result = []
        row_index = 2
        for row in self.sheet.iter_rows(min_row=row_index, max_col=1):
            for cell in row:
                if cell.value is None:  # 遇到空行(id为空)，则后面不再更新
                    break
                row_result = []
                for column in WEB_SHOW_COLUMNS:
                    cell = self.sheet.cell(row=row_index, column=column)
                    # value = reader.get_cell_value(cell.coordinate, FUND_SHEET_NAME)
                    value = get_cell_value(self.sheet, cell)
                    row_result.append(value)
                    if column == 1 or column == 2:  # stock index id cell has stock link, stock name cell has fund link
                        stock_id = self.sheet.cell(row=row_index, column=1).value
                        link = self.get_stock_hyperlink(stock_id, cell.row, column)
                        row_result.append(link)
                result.append(row_result)
                row_index = row_index + 1
            else:
                # Continue if the inner loop wasn't broken.
                continue
            # Inner loop was broken, break the outer.
            break
        return result

    def get_stock_hyperlink(self, stock_id, row, column):
        if column == 1:  # return stock link
            link_value = self.sheet.cell(row=row, column=STOCK_LINK_COLUMN).value
            if str(link_value).startswith('http') or str(link_value).startswith('no_stock_link'):
                # 如果明确没有可访问的stock链接，则Excel里面stock链接列以no_stock_link开头，直接返回，避免后面尝试哪个链接可用，很费时间
                return link_value
            return get_a_stock_link(stock_id)

        elif column == 2:  # return funds link, 指数宝链接
            return get_zhishubao_link(stock_id)

    def save_table(self, data):
        # 先把所有行第二列（名称）改为delete，然后根据web传来的数据覆盖excel内容，最后把名称为delete的删掉
        mark_name_delete(self.sheet, 2, 2)
        for row_data in data:   # for each row data
            row_index = find_value_row_index(self.sheet, 2, 1, row_data[0])
            # print('row data:', row_data[0], row_data[1], ", row_index=", row_index)
            del row_data[1]  # 删除 stock link, 下标从0开始
            del row_data[2]  # 删除 zhishubao link, 下标从0开始 -- 本来下标是3，但是前面移除了一个，变成2
            if row_index is None:   # 原来不存在添加行
                insert_row(self.sheet, 2, 1, row_data, WEB_SHOW_COLUMNS)
            else:
                set_row_data(self.sheet, row_index, row_data, 2, WEB_SHOW_COLUMNS)
        delete_rows(self.sheet, 2, 2)


# 这个请求多了很耗时，避免它运行的两个方法：
# 1) 如果确定没有stock link，则指数链接列写'no_stock_link'
# 2）如果链接不能根据stock id遵循一定规则变化过来（.SH,.SZ)，则直接在指数链接写上对应的http开头的链接。
def is_link_valid(url):
    import requests
    try:
        request = requests.get(url)
        http_status_code = request.status_code
        if http_status_code == 200:
            return True
    except requests.exceptions.HTTPError as e:
        pass
    return False


# stock_id is NNNNNN
def get_valid_stock_link(stock_id):
    for stock in ['sh'+stock_id, 'sz'+stock_id]:
        link = get_real_stock_link(stock)
        if is_link_valid(link):
            return link
    return DEFAULT_STOCK_LINK


def get_real_stock_link(stock):
    return "https://finance.sina.com.cn/realstock/company/" + stock + "/nc.shtml"


# stock_id is  NNNNNN.SZ / SH / CSI / SW
def get_a_stock_link(stock_id):
    stock_id_parts = str(stock_id).split('.')
    if len(stock_id_parts) < 2:
        return DEFAULT_STOCK_LINK
    if stock_id_parts[1] in ['SH', 'SZ']:
        stock = stock_id_parts[1].lower() + stock_id_parts[0]
        return get_real_stock_link(stock)

    return get_valid_stock_link(stock_id_parts[0])


def get_zhishubao_real_link(code):
    return "https://zhishubao.1234567.com.cn/home/detail?code=" + str(code)


def get_zhishubao_link(stock_id):
    stock_id_parts = str(stock_id).split('.')
    if len(stock_id_parts) < 2:
        return get_zhishubao_real_link(stock_id)
    if stock_id_parts[1] in ['SH', 'SZ', 'CSI', 'MI', 'SI', 'CI', 'CIC']:
        return get_zhishubao_real_link(stock_id_parts[0])
    if stock_id_parts[1] == 'HSI':
        return get_zhishubao_real_link(stock_id_parts[1])


if __name__ == '__main__':
    wb = openpyxl.load_workbook('example_filetest.xlsx')
    sheet = StockIndexSheet(wb)
    print(sheet.get_current_price("NDX"))
    # wb.save('example_filetest2.xlsx')
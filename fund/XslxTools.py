# -*- coding: utf-8 -*-

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font
from openpyxl.formula.translate import Translator
from openpyxl.cell.cell import TYPE_FORMULA
from openpyxl.styles.numbers import FORMAT_NUMBER_00
import re


MARK_AS_DELETE = "delete"

BUY_OFFSET = 100    # 已买基金在level的位置加上offset


# 设置positive, negative 条件格式
def set_p_n_condition(sheet, cell):
    red_text = Font(color="AA110D")
    red_fill = PatternFill(start_color='AA110D', end_color='FFC7CE', fill_type='solid')
    sheet.conditional_formatting.add(cell.coordinate, CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, font=red_text, fill=red_fill))
    green_text = Font(color="006100")
    green_fill = PatternFill(start_color='006100', end_color='C6EFCE', fill_type='solid')
    sheet.conditional_formatting.add(cell.coordinate, CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=green_text, fill=green_fill))


def set_double_p_n_condition(sheet, cell):
    set_p_n_condition(sheet, cell)
    cell.number_format = FORMAT_NUMBER_00


def get_formula_type(cell):
    # return cell.TYPE_FORMULA   # 直接在pycharm中运行，要使用这个（可能是openpyxl的新版本)
    return TYPE_FORMULA    # 在flask中运行，要用这个


def get_cell_value(sheet, cell, default=''):
    import logging
    _logger = logging.getLogger('werkzeug')

    if cell.value is None:
        return default
    if cell.data_type is not get_formula_type(cell):
        return cell.value

    #计算公式
    formula = cell.value[1:]  # 去掉等号
    match = re.match(".*?([A-Z]\d+).*?([A-Z]\d+).*?([A-Z]\d+).*", formula)
    if match is None:  # 遇到不同的公式
        return cell.value
    for address in match.groups():
        if sheet[address].value is None:
            return default
        formula = formula.replace(address, str(sheet[address].value))
    try:
        value = eval(formula)
    except TypeError as err:
        _logger.error("eval formula fail: " + str(formula) + " TypeError: " + str(err))
        return default
    except SyntaxError as err:
        _logger.error("eval formula fail: " + str(formula) + " SyntaxError: " + str(err))
        return default
    return value


def set_cell_value(sheet, cell, value, start_row):
    if cell.data_type is get_formula_type(cell):
        return    # 目标单元格已经有公式，不覆盖它
    else:
        start_cell = sheet.cell(row=start_row, column=cell.col_idx)
        if start_cell.data_type is get_formula_type(cell):  # 新插入单元格还没有公式，但是该列应该有公式
            # 从start_row对应位置拷贝公式到当前单元格
            cell.value = Translator(start_cell.value, origin=start_cell.coordinate).translate_formula(cell.coordinate)
            return
    cell.value = value


def find_value_row_index(sheet, start_row, col, value):
    for row in sheet.iter_rows(min_row=start_row, min_col=col, max_col=col):
        for cell in row:
            if str(cell.value) == str(value):
                return cell.row
    return None


# 考虑删除一行时，把改行后面的每一行的公式都做个调整
def process_formula_for_rows(sheet, start_row):
    for row in sheet.iter_rows(min_row=start_row, min_col=1):
        for cell in row:
            if cell.data_type is get_formula_type(cell):
                new_position = sheet.cell(row=cell.row - 1, column=cell.col_idx)
                cell.value = Translator(cell.value, origin=cell.coordinate).translate_formula(
                    new_position.coordinate)


# 把col列为value的行都删除
def delete_rows(sheet, start_row, col):
    while True:
        row_index = find_value_row_index(sheet, start_row, col, MARK_AS_DELETE)
        if row_index is None:
            break
        process_formula_for_rows(sheet, row_index + 1)
        #sheet.delete_rows(row_index, 1, translate=True)   # 删除一行，并且自动对公式进行translate处理， 这个还不支持
        # https://bitbucket.org/openpyxl/openpyxl/issues?status=new&status=open
        sheet.delete_rows(row_index, 1)


def find_last_row_index(sheet, start_row, col):
    row_index = start_row
    for row in sheet.iter_rows(min_row=start_row, min_col=col, max_col=col):
        for cell in row:
            if cell.value is None or str(cell.value).strip() == '':    # 找到None或该单元格内容为空，认为到达最后一行
                return cell.row
        row_index += 1
    return row_index


def set_row_data(sheet, row_index, data, start_row, columns=None):
    col_number = max(sheet.max_column, len(data))   # 如果网页插入列，则len(data)比原有sheet max_column大，应该取len(data)，否则会丢失列
    col_range = get_index_range(col_number)
    if columns is not None:
        col_range = columns
    array_index = 0
    for column in col_range:
        cell = sheet.cell(row=row_index, column=column)
        if str(cell.value) == MARK_AS_DELETE and data[array_index] is None:
            pass  # 如果标记为删除，且web传来的对应数据为None（即web上用户已删除），则不覆盖删除标记，便于最终删除它
        else:
            set_cell_value(sheet, cell, data[array_index], start_row)
        array_index += 1


# 在最后一行后面插入一行
def insert_row(sheet, start_row, id_col, data, columns=None):
    last_row_index = find_last_row_index(sheet, start_row, id_col)
    set_row_data(sheet, last_row_index, data, start_row, columns)


# excel from 1
def get_index_range(max_index):
    return range(1, max_index+1, 1)


# 全空白字符，不要转换为float
def str_to_float(value):
    if value is None:
        return None
    if isinstance(value,str):
        no_blank_value = value.strip()
        if no_blank_value == '':  # 空白字符串
            return None
        return float(no_blank_value)
    return float(value)   # 其它情况多半都是数字，直接返回


def is_value_empty(value):
    if value is None:
        return True
    if isinstance(value,str):
        no_blank_value = value.strip()
        if no_blank_value == '':  # 空白字符串
            return True
    return False


def str_to_int(value, default=0):
    if is_value_empty(value):
        return default
    if isinstance(value,str):
        no_blank_value = value.strip()
        return int(no_blank_value)
    return value


def calc_change_ratio(old_value, new_value):
    return (new_value - old_value)/old_value * 100


def update_focus_level(sheet, focus_level_column, row, current_fund_buy):
    focus_level_cell = sheet.cell(column=focus_level_column, row=row)
    old_value = get_cell_value(sheet, focus_level_cell, 0)
    old_value = str_to_int(old_value)
    new_value = old_value
    if current_fund_buy:
        if old_value < BUY_OFFSET:
            new_value = old_value + BUY_OFFSET
    else:
        if old_value >= BUY_OFFSET:
            new_value = old_value - BUY_OFFSET
    focus_level_cell.value = new_value


def mark_name_delete(sheet, start_row, col):
    row_index = start_row
    for row in sheet.iter_rows(min_row=row_index, min_col=col, max_col=col):
        for cell in row:
            cell.value = MARK_AS_DELETE
            row_index = row_index + 1